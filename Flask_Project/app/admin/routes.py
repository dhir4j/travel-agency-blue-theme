from flask import Blueprint, request, jsonify, send_file
from werkzeug.security import generate_password_hash
from app.models import User, Booking, Invoice
from app.extensions import db
from datetime import datetime, timedelta
from sqlalchemy import func
import io
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

admin_bp = Blueprint('admin', __name__, url_prefix="/api/admin")

# ========== User Management ==========

@admin_bp.route("/users", methods=["GET"])
def get_all_users():
    """Get all users with pagination"""
    limit = request.args.get("limit", type=int, default=50)
    offset = request.args.get("offset", type=int, default=0)
    search = request.args.get("search", "")

    query = User.query

    # Search filter
    if search:
        search_filter = f"%{search}%"
        query = query.filter(
            (User.email.ilike(search_filter)) |
            (User.first_name.ilike(search_filter)) |
            (User.last_name.ilike(search_filter))
        )

    total = query.count()
    users = query.order_by(User.created_at.desc()).limit(limit).offset(offset).all()

    return jsonify({
        "users": [user.to_dict() for user in users],
        "total": total,
        "limit": limit,
        "offset": offset
    }), 200


@admin_bp.route("/users/<int:user_id>", methods=["DELETE"])
def delete_user(user_id):
    """Delete a user"""
    user = User.query.get(user_id)

    if not user:
        return jsonify({"error": "User not found"}), 404

    if user.is_admin:
        return jsonify({"error": "Cannot delete admin user"}), 403

    db.session.delete(user)
    db.session.commit()

    return jsonify({"message": "User deleted successfully"}), 200


@admin_bp.route("/users/<int:user_id>/toggle-admin", methods=["PUT"])
def toggle_admin(user_id):
    """Toggle admin status for a user"""
    user = User.query.get(user_id)

    if not user:
        return jsonify({"error": "User not found"}), 404

    user.is_admin = not user.is_admin
    db.session.commit()

    return jsonify({
        "message": f"User admin status updated to {user.is_admin}",
        "user": user.to_dict()
    }), 200


# ========== Booking Management ==========

@admin_bp.route("/bookings", methods=["GET"])
def get_all_admin_bookings():
    """Get all bookings for admin dashboard"""
    limit = request.args.get("limit", type=int, default=50)
    offset = request.args.get("offset", type=int, default=0)
    status = request.args.get("status")
    payment_status = request.args.get("payment_status")
    search = request.args.get("search", "")

    query = Booking.query

    # Filters
    if status:
        query = query.filter_by(status=status)
    if payment_status:
        query = query.filter_by(payment_status=payment_status)
    if search:
        search_filter = f"%{search}%"
        query = query.filter(
            (Booking.booking_id_str.ilike(search_filter)) |
            (Booking.user_email.ilike(search_filter)) |
            (Booking.destination.ilike(search_filter))
        )

    total = query.count()
    bookings = query.order_by(Booking.booking_date.desc()).limit(limit).offset(offset).all()

    return jsonify({
        "bookings": [booking.to_dict() for booking in bookings],
        "total": total,
        "limit": limit,
        "offset": offset
    }), 200


@admin_bp.route("/bookings/<int:booking_id>/status", methods=["PUT"])
def update_booking_status(booking_id):
    """Update booking status"""
    booking = Booking.query.get(booking_id)

    if not booking:
        return jsonify({"error": "Booking not found"}), 404

    data = request.get_json()
    new_status = data.get("status")
    new_payment_status = data.get("payment_status")

    if new_status:
        booking.status = new_status
    if new_payment_status:
        booking.payment_status = new_payment_status

        # Update invoice status if payment status changed
        if booking.invoice:
            booking.invoice.status = new_payment_status

    db.session.commit()

    return jsonify({
        "message": "Booking status updated",
        "booking": booking.to_dict()
    }), 200


# ========== Dashboard Statistics ==========

@admin_bp.route("/stats/dashboard", methods=["GET"])
def get_dashboard_stats():
    """Get dashboard statistics"""

    # Total users
    total_users = User.query.count()
    new_users_this_month = User.query.filter(
        User.created_at >= datetime.utcnow().replace(day=1, hour=0, minute=0, second=0)
    ).count()

    # Total bookings
    total_bookings = Booking.query.count()
    pending_bookings = Booking.query.filter_by(status="Pending").count()
    confirmed_bookings = Booking.query.filter_by(status="Confirmed").count()

    # Revenue stats
    total_revenue = db.session.query(func.sum(Booking.final_amount)).scalar() or 0
    paid_revenue = db.session.query(func.sum(Invoice.paid_amount)).scalar() or 0
    pending_revenue = db.session.query(func.sum(Invoice.balance_due)).scalar() or 0

    # This month stats
    first_day_of_month = datetime.utcnow().replace(day=1, hour=0, minute=0, second=0)
    this_month_bookings = Booking.query.filter(Booking.booking_date >= first_day_of_month).count()
    this_month_revenue = db.session.query(func.sum(Booking.final_amount)).filter(
        Booking.booking_date >= first_day_of_month
    ).scalar() or 0

    # Recent bookings
    recent_bookings = Booking.query.order_by(Booking.booking_date.desc()).limit(5).all()

    return jsonify({
        "users": {
            "total": total_users,
            "new_this_month": new_users_this_month
        },
        "bookings": {
            "total": total_bookings,
            "pending": pending_bookings,
            "confirmed": confirmed_bookings,
            "this_month": this_month_bookings
        },
        "revenue": {
            "total": float(total_revenue),
            "paid": float(paid_revenue),
            "pending": float(pending_revenue),
            "this_month": float(this_month_revenue)
        },
        "recent_bookings": [booking.to_dict() for booking in recent_bookings]
    }), 200


@admin_bp.route("/stats/analytics", methods=["GET"])
def get_analytics():
    """Get analytics data for charts"""
    days = request.args.get("days", type=int, default=30)
    start_date = datetime.utcnow() - timedelta(days=days)

    # Daily bookings
    daily_bookings = db.session.query(
        func.date(Booking.booking_date).label('date'),
        func.count(Booking.id).label('count'),
        func.sum(Booking.final_amount).label('revenue')
    ).filter(
        Booking.booking_date >= start_date
    ).group_by(
        func.date(Booking.booking_date)
    ).all()

    # Bookings by status
    status_breakdown = db.session.query(
        Booking.status,
        func.count(Booking.id).label('count')
    ).group_by(Booking.status).all()

    # Bookings by package type
    package_breakdown = db.session.query(
        Booking.package_type,
        func.count(Booking.id).label('count'),
        func.sum(Booking.final_amount).label('revenue')
    ).group_by(Booking.package_type).all()

    return jsonify({
        "daily_bookings": [
            {
                "date": str(item.date),
                "count": item.count,
                "revenue": float(item.revenue or 0)
            }
            for item in daily_bookings
        ],
        "status_breakdown": [
            {"status": item.status, "count": item.count}
            for item in status_breakdown
        ],
        "package_breakdown": [
            {
                "package_type": item.package_type,
                "count": item.count,
                "revenue": float(item.revenue or 0)
            }
            for item in package_breakdown
        ]
    }), 200


# ========== Reports & Downloads ==========

@admin_bp.route("/reports/bookings/csv", methods=["GET"])
def download_bookings_csv():
    """Download bookings report as CSV"""
    status = request.args.get("status")
    start_date = request.args.get("start_date")
    end_date = request.args.get("end_date")

    query = Booking.query

    # Apply filters
    if status:
        query = query.filter_by(status=status)
    if start_date:
        query = query.filter(Booking.booking_date >= datetime.fromisoformat(start_date))
    if end_date:
        query = query.filter(Booking.booking_date <= datetime.fromisoformat(end_date))

    bookings = query.order_by(Booking.booking_date.desc()).all()

    # Create CSV in memory
    output = io.StringIO()
    writer = csv.writer(output)

    # Write header
    writer.writerow([
        'Booking ID', 'User Email', 'Package Name', 'Package Type', 'Destination',
        'Travel Date', 'Return Date', 'Adults', 'Children', 'Total Amount',
        'Tax', 'Discount', 'Final Amount', 'Status', 'Payment Status',
        'Booking Date'
    ])

    # Write data
    for booking in bookings:
        writer.writerow([
            booking.booking_id_str,
            booking.user_email,
            booking.package_name,
            booking.package_type,
            booking.destination,
            booking.travel_date.isoformat() if booking.travel_date else '',
            booking.return_date.isoformat() if booking.return_date else '',
            booking.num_adults,
            booking.num_children,
            float(booking.total_amount),
            float(booking.tax_amount),
            float(booking.discount_amount),
            float(booking.final_amount),
            booking.status,
            booking.payment_status,
            booking.booking_date.isoformat() if booking.booking_date else ''
        ])

    # Prepare response
    output.seek(0)
    return send_file(
        io.BytesIO(output.getvalue().encode()),
        mimetype='text/csv',
        as_attachment=True,
        download_name=f'bookings_report_{datetime.now().strftime("%Y%m%d")}.csv'
    )


@admin_bp.route("/reports/bookings/excel", methods=["GET"])
def download_bookings_excel():
    """Download bookings report as Excel"""
    status = request.args.get("status")
    start_date = request.args.get("start_date")
    end_date = request.args.get("end_date")

    query = Booking.query

    # Apply filters
    if status:
        query = query.filter_by(status=status)
    if start_date:
        query = query.filter(Booking.booking_date >= datetime.fromisoformat(start_date))
    if end_date:
        query = query.filter(Booking.booking_date <= datetime.fromisoformat(end_date))

    bookings = query.order_by(Booking.booking_date.desc()).all()

    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Bookings Report"

    # Style header
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4361EE", end_color="4361EE", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    # Write header
    headers = [
        'Booking ID', 'User Email', 'Package Name', 'Package Type', 'Destination',
        'Travel Date', 'Return Date', 'Adults', 'Children', 'Total Amount',
        'Tax', 'Discount', 'Final Amount', 'Status', 'Payment Status',
        'Booking Date'
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # Write data
    for row, booking in enumerate(bookings, 2):
        ws.cell(row=row, column=1, value=booking.booking_id_str)
        ws.cell(row=row, column=2, value=booking.user_email)
        ws.cell(row=row, column=3, value=booking.package_name)
        ws.cell(row=row, column=4, value=booking.package_type)
        ws.cell(row=row, column=5, value=booking.destination)
        ws.cell(row=row, column=6, value=booking.travel_date.isoformat() if booking.travel_date else '')
        ws.cell(row=row, column=7, value=booking.return_date.isoformat() if booking.return_date else '')
        ws.cell(row=row, column=8, value=booking.num_adults)
        ws.cell(row=row, column=9, value=booking.num_children)
        ws.cell(row=row, column=10, value=float(booking.total_amount))
        ws.cell(row=row, column=11, value=float(booking.tax_amount))
        ws.cell(row=row, column=12, value=float(booking.discount_amount))
        ws.cell(row=row, column=13, value=float(booking.final_amount))
        ws.cell(row=row, column=14, value=booking.status)
        ws.cell(row=row, column=15, value=booking.payment_status)
        ws.cell(row=row, column=16, value=booking.booking_date.isoformat() if booking.booking_date else '')

    # Auto-size columns
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'bookings_report_{datetime.now().strftime("%Y%m%d")}.xlsx'
    )


@admin_bp.route("/invoices", methods=["GET"])
def get_all_invoices():
    """Get all invoices"""
    limit = request.args.get("limit", type=int, default=50)
    offset = request.args.get("offset", type=int, default=0)
    status = request.args.get("status")

    query = Invoice.query

    if status:
        query = query.filter_by(status=status)

    total = query.count()
    invoices = query.order_by(Invoice.invoice_date.desc()).limit(limit).offset(offset).all()

    # Get booking details for each invoice
    invoice_data = []
    for invoice in invoices:
        inv_dict = invoice.to_dict()
        inv_dict['booking'] = invoice.booking.to_dict() if invoice.booking else None
        invoice_data.append(inv_dict)

    return jsonify({
        "invoices": invoice_data,
        "total": total,
        "limit": limit,
        "offset": offset
    }), 200


@admin_bp.route("/invoices/<int:invoice_id>", methods=["GET"])
def get_invoice(invoice_id):
    """Get a specific invoice"""
    invoice = Invoice.query.get(invoice_id)

    if not invoice:
        return jsonify({"error": "Invoice not found"}), 404

    invoice_dict = invoice.to_dict()
    invoice_dict['booking'] = invoice.booking.to_dict() if invoice.booking else None

    return jsonify({"invoice": invoice_dict}), 200
